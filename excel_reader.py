import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional
import os
import re

def safe_float(value):
    """Convert value to float safely, handling commas and formatting (Indonesian)"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace('Rp', '').replace('Rp.', '').replace('Rp ', '').strip()
        # Handle Indonesian format: 1.234.567,89 -> 1234567.89  or 280,000 -> 280000
        # If contains ',' as decimal separator, normalize
        if ',' in cleaned and '.' in cleaned:
            # 1.234,50
            cleaned = cleaned.replace('.', '').replace(',', '.')
        elif ',' in cleaned:
            # Check if comma is thousands separator (3 digits after comma) or decimal
            parts = cleaned.split(',')
            if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit() and parts[0].replace('-','').isdigit():
                # Likely thousands separator: 280,000
                cleaned = cleaned.replace(',', '')
            elif len(parts) == 2 and len(parts[1]) <= 2:
                # Decimal: 1,5
                cleaned = cleaned.replace(',', '.')
            else:
                # Multiple commas or other -> treat as thousands
                cleaned = cleaned.replace(',', '').replace('.', '')
        else:
            # No comma, just remove dots (thousands)
            cleaned = cleaned.replace('.', '')
        cleaned = cleaned.strip()
        if cleaned == '' or cleaned == '-':
            return None
        try:
            return float(cleaned)
        except:
            return None
    return None

class ExcelReader:
    """Class untuk membaca file Excel RAB"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.wb = None
        self.wb_data = None
        self.ws = None
        self.ws_data = None
        self.df = None
        
    def load_workbook(self) -> bool:
        """Load workbook - 2 versi: formula dan data"""
        try:
            self.wb = load_workbook(self.file_path, data_only=False)
            self.wb_data = load_workbook(self.file_path, data_only=True)
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def get_sheet_names(self) -> List[str]:
        """Mendapatkan nama-nama sheet"""
        if self.wb is None:
            if not self.load_workbook():
                return []
        return self.wb.sheetnames
    
    def select_sheet(self, sheet_name: str) -> bool:
        """Pilih sheet yang akan dianalisis"""
        try:
            if self.wb is None:
                if not self.load_workbook():
                    return False
            self.ws = self.wb[sheet_name]
            self.ws_data = self.wb_data[sheet_name]
            return True
        except Exception as e:
            print(f"Error selecting sheet: {e}")
            return False
    
    def find_header_row(self) -> Optional[int]:
        """Mencari baris header (kolom QTY, Unit Price, TOTAL)"""
        if self.ws is None:
            return None
        
        for row in range(1, self.ws.max_row + 1):
            row_values = []
            for col in range(1, self.ws.max_column + 1):
                cell_value = self.ws.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value).upper().strip())
            
            has_qty = any('QTY' in val or 'QUANTITY' in val or 'JUMLAH' in val for val in row_values)
            has_total = any('TOTAL' in val or 'HARGA TOTAL' in val for val in row_values)
            
            if has_qty and has_total:
                return row
        return None
    
    def find_data_columns(self, header_row: int) -> Dict[str, int]:
        """Mencari posisi kolom data"""
        columns = {
            'item_name': None,
            'qty': None,
            'harga_awal': None,
            'mark_up': None,
            'unit_price': None,
            'total': None
        }
        
        header_values = {}
        for col in range(1, self.ws.max_column + 1):
            cell_value = self.ws.cell(row=header_row, column=col).value
            if cell_value:
                header_values[str(cell_value).upper().strip()] = col
        
        for key, value in header_values.items():
            if 'QTY' in key or 'QUANTITY' in key or 'JUMLAH' in key:
                columns['qty'] = value
            elif 'UNIT PRICE' in key or 'HARGA SATUAN' in key:
                columns['unit_price'] = value
            elif 'TOTAL' in key:
                columns['total'] = value
            elif 'HARGA AWAL' in key or 'PRICE' in key:
                columns['harga_awal'] = value
            elif 'MARK UP' in key or 'MARKUP' in key:
                columns['mark_up'] = value
            elif 'ITEM' in key or 'NAME' in key or 'NAMA' in key or 'DESKRIPSI' in key or 'DESCRIPTION' in key or 'KETERANGAN' in key or 'BARANG' in key:
                columns['item_name'] = value
        
        # Fallback: infer columns from numeric patterns if header detection failed
        if columns['qty'] is None or columns['unit_price'] is None or columns['total'] is None:
            numeric_cols = []
            for col in range(1, self.ws.max_column + 1):
                numeric_count = 0
                for r in range(header_row + 1, min(header_row + 8, self.ws.max_row + 1)):
                    v = self.ws_data.cell(row=r, column=col).value
                    if v is None:
                        v = self.ws.cell(row=r, column=col).value
                    if isinstance(v, (int, float)) and v > 0:
                        numeric_count += 1
                    elif isinstance(v, str) and v.strip().replace(',','').replace('.','').strip().isdigit():
                        # string numeric like "280,000"
                        numeric_count += 1
                if numeric_count >= 2:
                    numeric_cols.append(col)
            numeric_cols = sorted(set(numeric_cols))
            # Heuristic: qty is smallest/col leftmost among numeric, total is rightmost
            if columns['qty'] is None and numeric_cols:
                columns['qty'] = numeric_cols[0]
            if columns['total'] is None and numeric_cols:
                columns['total'] = numeric_cols[-1]
            if columns['unit_price'] is None and len(numeric_cols) >= 2:
                # pick middle numeric col
                if len(numeric_cols) >= 3:
                    columns['unit_price'] = numeric_cols[-2]
                else:
                    columns['unit_price'] = numeric_cols[1] if numeric_cols[1] != columns['total'] else numeric_cols[0]
        
        if columns['qty'] is None:
            columns['qty'] = 6
        if columns['unit_price'] is None:
            columns['unit_price'] = 9
        if columns['total'] is None:
            columns['total'] = 10
        
        if columns['item_name'] is None:
            for col in range(1, columns['qty']):
                has_text = False
                for row in range(header_row + 1, min(header_row + 5, self.ws.max_row + 1)):
                    cell_value = self.ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value) > 2:
                        has_text = True
                        break
                if has_text:
                    columns['item_name'] = col
                    break
        
        return columns
    
    def _get_total_value(self, row: int, total_col: int) -> Any:
        """Ambil nilai dari kolom total"""
        total_cell_data = self.ws_data.cell(row=row, column=total_col)
        total_cell_formula = self.ws.cell(row=row, column=total_col)
        
        raw_val = None
        if total_cell_data.value is not None:
            raw_val = total_cell_data.value
        elif total_cell_formula.value is not None:
            if isinstance(total_cell_formula.value, str) and total_cell_formula.value.startswith('='):
                return None
            else:
                raw_val = total_cell_formula.value
        
        if raw_val is not None:
            converted = safe_float(raw_val)
            return converted if converted is not None else raw_val
        return None
    
    def _normalize_label(self, text: str) -> str:
        """Normalisasi label: ringkas spasi, hapus .:()- dan uppercase. Untuk toleran typo."""
        s = text.upper().strip()
        # Hapus separator umum yang sering bikin typo spasi
        s = s.replace('.', ' ').replace(':', ' ').replace('(', ' ').replace(')', ' ').replace('-', ' ').replace('/', ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def _detect_section_letter(self, text: str) -> Optional[str]:
        """Deteksi huruf section dari text seperti 'Total A', 'Jumlah B' — toleran typo. TOTAL (A+B[+C]) bukan section."""
        norm = self._normalize_label(text)
        if 'GRAND TOTAL' in norm or 'GRANDTOTAL' in norm.replace(' ', ''):
            return 'GRAND'
        if '+' in norm and re.search(r'TOTAL\s*\(.*\+', norm):
            return None
        norm_jml = re.sub(r'\bJML\b', 'JUMLAH', norm)
        m2 = re.search(r'(?:TOTAL|JUMLAH|SUBTOTAL)[\s\.]*([A-Z])\b', norm_jml)
        if m2:
            if '+' in norm_jml:
                return None
            return m2.group(1)
        m3 = re.search(r'(?:TOTAL|JUMLAH|SUBTOTAL)([A-Z])\b', norm_jml.replace(' ', ''))
        if m3:
            if '+' in norm_jml:
                return None
            return m3.group(1)
        try:
            from rapidfuzz import fuzz
            for target in ['TOTAL', 'JUMLAH', 'SUBTOTAL']:
                token = norm_jml.split()[0] if norm_jml.split() else ''
                if token and fuzz.ratio(token, target) >= 85 and re.search(r'[A-Z]\b', norm_jml):
                    ml = re.search(r'([A-Z])\b', norm_jml)
                    if ml:
                        return ml.group(1)
        except ImportError:
            pass
        return None

    def _detect_row_type(self, cell_str: str) -> str:
        """Deteksi jenis baris: section_header, subtotal, ppn, discount, total, grand_total — toleran typo."""
        norm = self._normalize_label(cell_str)
        norm_jml = re.sub(r'\bJML\b', 'JUMLAH', norm)
        
        # GRAND TOTAL — jangan makan TOTAL (A+B) yang sebelum PPN; hanya GRAND TOTAL asli
        if 'GRAND TOTAL' in norm or 'GRANDTOTAL' in norm.replace(' ', ''):
            return 'grand_total'
        # Fuzzy GRAND TOTAL
        try:
            from rapidfuzz import fuzz
            if fuzz.partial_ratio(norm, 'GRAND TOTAL') >= 85 and 'TOTAL' in norm:
                return 'grand_total'
        except ImportError:
            pass
        
        # Section header standalone (A, B, C) — hanya kalau benar-benar 1 huruf
        # Jangan match satuan 'm' yang sudah di-filter di loop (col==1 + cek qty)
        stripped = norm.strip()
        if len(stripped) == 1 and stripped.isalpha() and stripped in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            return 'section_header'
        
        # PPN/Tax — toleran, tapi JANGAN match footer "Penawaran sudah termasuk PPN"
        # Footer seperti "2. Penawaran sudah termasuk PPN 11%" bukan baris ringkasan -> skip via pengecekan awal
        if cell_str.strip().startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.')):
            # Kemungkinan notes/footer, bukan ringkasan — tapi tetap cek jika baris itu BENAR-BENAR berlabel "PPN" saja
            # Untuk aman, jangan deteksi PPN jika teks > 30 char dan mengandung "PENAWARAN" / "PEMBAYARA" / "SYARAT"
            upper_raw = cell_str.upper()
            if len(cell_str) > 30 and ('PENAWARAN' in upper_raw or 'PEMBAYARA' in upper_raw or 'SYARAT' in upper_raw or 'KETENTUAN' in upper_raw):
                pass  # skip deteksi PPN untuk footer panjang
            else:
                ppn_norm = norm.replace(' ', '')
                if 'PPN' in norm or 'PPN' in ppn_norm or 'PAJAK' in norm or norm_jml.replace(' ', '') == 'TAX':
                    if re.search(r'\bPPN\b', norm) or 'PPN' in ppn_norm or re.search(r'\bPAJAK\b', norm) or re.search(r'\bTAX\b', norm):
                        return 'ppn'
        else:
            ppn_norm = norm.replace(' ', '')
            if 'PPN' in norm or 'PPN' in ppn_norm or 'PAJAK' in norm or norm_jml.replace(' ', '') == 'TAX':
                if re.search(r'\bPPN\b', norm) or 'PPN' in ppn_norm or re.search(r'\bPAJAK\b', norm) or re.search(r'\bTAX\b', norm):
                    return 'ppn'
        try:
            from rapidfuzz import fuzz
            for tok in norm.split():
                if fuzz.ratio(tok, 'PPN') >= 80 or fuzz.ratio(tok, 'PAJAK') >= 80:
                    # Hindari footer panjang
                    if len(cell_str) > 30 and 'PENAWARAN' in cell_str.upper():
                        continue
                    return 'ppn'
        except ImportError:
            pass
        if 'PPN' in cell_str.upper() or 'TAX' in cell_str.upper() or 'PAJAK' in cell_str.upper():
            if len(cell_str) > 35 and 'PENAWARAN' in cell_str.upper():
                pass
            else:
                return 'ppn'
        if 'PPN' in norm or 'PAJAK' in norm:
            if len(cell_str) > 35 and 'PENAWARAN' in cell_str.upper():
                pass
            else:
                return 'ppn'
        
        # Discount/Diskon
        if 'DISKON' in norm or 'DISCOUNT' in norm or 'POTONGAN' in norm:
            return 'discount'
        try:
            from rapidfuzz import fuzz
            for tok in norm.split():
                if fuzz.ratio(tok, 'DISKON') >= 85 or fuzz.ratio(tok, 'DISCOUNT') >= 85:
                    return 'discount'
        except ImportError:
            pass
        
        # TOTAL (A+B[+C]) — normalisasi hapus (), jadi TOTAL A+B — deteksi via PLUS
        if 'TOTAL' in norm and '+' in norm:
            if norm.strip() != 'GRAND TOTAL' and norm.replace(' ','') != 'GRANDTOTAL':
                return 'jumlah_global'
        if 'TOTAL' in norm_jml and '+' in norm_jml:
            return 'jumlah_global'
        norm_for_global = norm_jml.strip()
        if 'GRAND TOTAL' in norm or 'GRANDTOTAL' in norm.replace(' ', ''):
            return 'grand_total'
        if norm_for_global in ('JUMLAH', 'TOTAL', 'SUBTOTAL', 'SUB TOTAL', 'JUMLAH SEBELUM PPN', 'TOTAL SEBELUM PPN', 'JUMLAH TOTAL', 'SUBTOTAL GLOBAL', 'TOTAL JUMLAH', 'JUMLAH GLOBAL'):
            return 'jumlah_global'
        if norm_for_global.startswith(('JUMLAH SEBELUM', 'TOTAL SEBELUM', 'SUBTOTAL SEBELUM')):
            return 'jumlah_global'
        if not self._detect_section_letter(cell_str):
            if norm_for_global in ('JUMLAH', 'TOTAL', 'SUB TOTAL', 'SUBTOTAL'):
                return 'jumlah_global'
            if norm_for_global.startswith('JUMLAH') and not re.search(r'JUMLAH[\s\.]*[A-Z]\b', norm_for_global):
                if len(norm_for_global) <= 25:
                    return 'jumlah_global'

        if 'TOTAL' in norm and '+' in norm:
            return 'unknown'
        if 'TOTAL' in norm_jml and '+' in norm_jml:
            return 'unknown'
        # Subtotal/Total section — toleran: TOTAL / JUMLAH / JML / SUBTOTAL / SUB TOTAL (huruf tunggal A/B/C saja)
        if ('TOTAL' in norm_jml or 'JUMLAH' in norm_jml or 'SUBTOTAL' in norm_jml or 'SUB TOTAL' in norm):
            # Hanya subtotal jika dengan huruf section TUNGGAL (Total A / Jumlah B), bukan gabungan (TOTAL (A+B))
            # Gabungan sudah di-handle sebagai jumlah_global di atas
            if len(norm.strip()) <= 2 and norm.strip().isalpha():
                return 'unknown'
            # Toleran: TOTAL A / Jumlah B saja; gabungan sudah jumlah_global
            if re.search(r'(?:TOTAL|JUMLAH|SUBTOTAL)[\s\.]*[A-Z]\b', norm_jml) and not re.search(r'\(.*\+', norm_jml):
                return 'subtotal'
            if norm_jml in ('TOTAL','JUMLAH','SUBTOTAL') or norm_jml.startswith('TOTAL ') or norm_jml.startswith('JUMLAH ') or 'SUBTOTAL' in norm_jml:
                # Generic TOTAL/JUMLAH tanpa huruf section -> tetap subtotal (single section)
                if not re.search(r'\(.*\+', norm_jml):
                    return 'subtotal'
            return 'unknown'
        # Fuzzy fallback untuk TOTAL/JUMLAH typo 1 huruf
        try:
            from rapidfuzz import fuzz
            first_tok = norm_jml.split()[0] if norm_jml.split() else ''
            if first_tok and (fuzz.ratio(first_tok, 'TOTAL') >= 85 or fuzz.ratio(first_tok, 'JUMLAH') >= 85 or fuzz.ratio(first_tok, 'SUBTOTAL') >= 85):
                return 'subtotal'
        except ImportError:
            pass
        
        return 'unknown'
    
    def read_data(self, sheet_name: str, overrides: Optional[Dict[str, Any]] = None, ai_overrides: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        """AI primary: jika ada ai_overrides (Gemini/Groq gratis), klasifikasi via AI dulu; fallback Value Intelligence."""
        # Untuk AI primary, kita jalankan deteksi awal via AI sebelum loop utama bila ada key
        # Tetapi tetap butuh classifications untuk AI — jadi AI akan dipanggil setelah loop klasifikasi awal
        # Simpan ai_overrides untuk dipakai di akhir sebelum return
        return self._read_data_impl(sheet_name, overrides, ai_overrides)

    def _read_data_impl(self, sheet_name: str, overrides: Optional[Dict[str, Any]] = None, ai_overrides: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        """Membaca data dari sheet dengan support multiple sections dan flexible fields.
        
        overrides (opsional, tanpa AI, 100% lokal):
            header_row: int | None — override baris header (1-indexed)
            qty_col, unit_price_col, total_col: int | None — huruf kolom A=1, B=2, ...
            ppn_mode: 'auto' | 'per_section' | 'combined' | 'single' | 'none'
            total_mode: 'auto' | 'per_section' | 'combined'  — combined = 1 Total global
        """
        overrides = overrides or {}
        result: Dict[str, Any] = {
            'sheet_name': sheet_name,
            'header_row': None,
            'columns': None,
            'items': [],
            'sections': {},
            'subtotal_row': None,
            'subtotal_value': None,
            'ppn_row': None,
            'ppn_value': None,
            'grand_total_row': None,
            'grand_total_value': None,
            'formulas': {},
            'overrides_applied': overrides
        }
        
        if not self.select_sheet(sheet_name):
            return result
        
        header_row = overrides.get('header_row') or self.find_header_row()
        if header_row is None:
            return result
        # Simpan header untuk adukan jika nanti TOTAL tidak kebaca (debug)
        result['header_row'] = header_row
        result['header_values_debug'] = [str(self.ws.cell(row=header_row, column=c).value) if self.ws.cell(row=header_row, column=c).value is not None else "" for c in range(1, min(16, self.ws.max_column+1))]
        
        columns = self.find_data_columns(header_row)
        # Apply column overrides
        if overrides.get('qty_col'):
            columns['qty'] = int(overrides['qty_col'])
        if overrides.get('unit_price_col'):
            columns['unit_price'] = int(overrides['unit_price_col'])
        if overrides.get('total_col'):
            columns['total'] = int(overrides['total_col'])
        result['columns'] = columns
        
        total_col = columns.get('total', 10)
        current_section = None
        pending_items = []
        section_order = []  # Track urutan section
        skipped_rows = []  # For debug
        ppn_candidates = []  # Kumpulkan semua baris PPN, klasifikasi setelah loop (lebih akurat)
        classifications = []  # Untuk DEBUG: log klasifikasi dengan normalisasi + flag typo
        
        for row in range(header_row + 1, self.ws.max_row + 1):
            is_summary_row = False
            
            # Cek semua kolom untuk mencari label
            row_classified = False
            for col in range(1, min(self.ws.max_column + 1, 15)):
                cell_value = self.ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    norm_dbg = self._normalize_label(cell_str)
                    row_type = self._detect_row_type(cell_str)
                    # Log klasifikasi ringkasan untuk DEBUG (skip item text biasa)
                    if row_type != 'unknown' and not row_classified:
                        is_fuzzy = False
                        try:
                            from rapidfuzz import fuzz
                            # Tandai typo jika mengandung JML/TOTAl typo tapi norm mengandunya
                            if norm_dbg != cell_str.upper().strip():
                                # Ada normalisasi (.:()- dihilangkan) -> kemungkinan typo minor
                                pass
                            # Fuzzy flag: jika first token fuzzy match tapi bukan exact
                            tok = norm_dbg.split()[0] if norm_dbg.split() else ''
                            for target in ['TOTAL','JUMLAH','SUBTOTAL','GRAND TOTAL','PPN','DISKON']:
                                if tok and fuzz.ratio(tok, target.split()[0]) >= 85 and tok != target.split()[0]:
                                    is_fuzzy = True
                                    break
                        except ImportError:
                            pass
                        if cell_str.upper().strip() != norm_dbg and re.search(r'[.:()\-/]', cell_str):
                            is_fuzzy = True
                        classifications.append({'row': row, 'raw': cell_str[:40], 'normalized': norm_dbg[:40], 'type': row_type, 'fuzzy': is_fuzzy})
                        row_classified = True
                    
                    if row_type == 'grand_total':
                        # Grand Total
                        grand_total_value = self._get_total_value(row, total_col)
                        if grand_total_value is not None:
                            result['grand_total_row'] = row
                            result['grand_total_value'] = grand_total_value
                            is_summary_row = True
                            break
                    
                    elif row_type == 'jumlah_global':
                        # TOTAL (A+B) sebelum PPN = jumlah_global — jangan timpa jika sudah ada PPN global
                        # Jika sudah ada grand_total (mis. PPN hanya di 1 bagian, grand sudah TOTAL A+B+PPN), jumlah_global adalah gabungan
                        val = self._get_total_value(row, total_col)
                        if val is None:
                            for c2 in range(self.ws.max_column, 0, -1):
                                try:
                                    v2 = self.ws_data.cell(row=row, column=c2).value
                                    if v2 is None:
                                        v2 = self.ws.cell(row=row, column=c2).value
                                    if v2 is None and c2 != 8:
                                        v2 = self.ws_data.cell(row=row, column=8).value
                                        if v2 is None:
                                            v2 = self.ws.cell(row=row, column=8).value
                                    sf = safe_float(v2)
                                    if sf is not None and sf > 0:
                                        val = sf
                                        break
                                except:
                                    pass
                        if val is not None:
                            result['jumlah_global_row'] = row
                            result['jumlah_global_excel'] = val
                            result['total_kategori_value'] = val
                            result['total_kategori_row'] = row
                            # Jangan timpa grand_total yang sudah benar
                        is_summary_row = True
                        break

                    elif row_type == 'section_header':
                        # Section header (A, B, C) — HANYA jika di kolom 1
                        # Bug sebelumnya: satuan "m" di kolom unit terdeteksi sebagai section "M"
                        if col != 1:
                            continue
                        # Pastikan bukan baris item yang kebetulan punya satuan "m"/"l"
                        # Jika baris punya qty numerik, ini bukan section header
                        _qty_check = None
                        _up_check = None
                        if columns.get('qty'):
                            _v = self.ws_data.cell(row=row, column=columns['qty']).value
                            if _v is None:
                                _v = self.ws.cell(row=row, column=columns['qty']).value
                            _qty_check = safe_float(_v)
                        if columns.get('unit_price'):
                            _v2 = self.ws_data.cell(row=row, column=columns['unit_price']).value
                            if _v2 is None:
                                _v2 = self.ws.cell(row=row, column=columns['unit_price']).value
                            _up_check = safe_float(_v2)
                        if _qty_check is not None or _up_check is not None:
                            continue
                        section_letter = cell_str.upper()
                        current_section = section_letter
                        
                        if section_letter not in result['sections']:
                            result['sections'][section_letter] = {
                                'items': [],
                                'subtotal_row': None,
                                'subtotal_value': None,
                                'ppn_row': None,
                                'ppn_value': None,
                                'discount_row': None,
                                'discount_value': None,
                                'total_row': None,
                                'total_value': None
                            }
                            section_order.append(section_letter)
                        
                        # Pindahkan pending items ke section ini
                        if pending_items:
                            result['sections'][section_letter]['items'] = pending_items
                            pending_items = []
                        
                        is_summary_row = True
                        break
                    
                    elif row_type == 'ppn':
                        # Kumpulkan dulu, klasifikasi setelah loop (saat semua subtotal sudah kebaca)
                        ppn_value = self._get_total_value(row, total_col)
                        if ppn_value is not None:
                            ppn_candidates.append({
                                'row': row, 'value': ppn_value, 'label': cell_str,
                                'current_section': current_section
                            })
                            is_summary_row = True
                            break
                    
                    elif row_type == 'discount':
                        # Diskon - simpan ke section yang aktif
                        discount_value = self._get_total_value(row, total_col)
                        if discount_value is not None:
                            if current_section and current_section in result['sections']:
                                result['sections'][current_section]['discount_value'] = discount_value
                                result['sections'][current_section]['discount_row'] = row
                            is_summary_row = True
                            break
                    
                    elif row_type == 'subtotal':
                        # TOTAL (A+B[+C]) sejenis subtotal karena + salah regex — paksa jumlah_global
                        if '+' in cell_str.upper() and 'TOTAL' in cell_str.upper():
                            val2 = self._get_total_value(row, total_col)
                            if val2 is None:
                                for c2 in [8, 6, 9, 10]:
                                    try:
                                        v2 = self.ws_data.cell(row=row, column=c2).value
                                        if v2 is None: v2 = self.ws.cell(row=row, column=c2).value
                                        sf = safe_float(v2)
                                        if sf is not None and sf > 0:
                                            val2 = sf; break
                                    except: pass
                            if val2 is not None:
                                result['jumlah_global_row'] = row
                                result['jumlah_global_excel'] = val2
                                result['total_kategori_value'] = val2
                                result['total_kategori_row'] = row
                            is_summary_row = True
                            break
                        total_value = self._get_total_value(row, total_col)
                        if total_value is not None:
                            section_letter = self._detect_section_letter(cell_str)
                            
                            if section_letter == 'GRAND':
                                result['grand_total_row'] = row
                                result['grand_total_value'] = total_value
                            elif section_letter:
                                # Total dengan label section (Total A, Jumlah B)
                                if section_letter not in result['sections']:
                                    result['sections'][section_letter] = {
                                        'items': [],
                                        'subtotal_row': None,
                                        'subtotal_value': None,
                                        'ppn_row': None,
                                        'ppn_value': None,
                                        'discount_row': None,
                                        'discount_value': None,
                                        'total_row': row,
                                        'total_value': total_value
                                    }
                                    section_order.append(section_letter)
                                else:
                                    # Check apakah ini "Total X" (sudah termasuk PPN) atau "Jumlah X" (belum PPN)
                                    cell_upper = cell_str.upper()
                                    if 'TOTAL' in cell_upper and 'JUMLAH' not in cell_upper:
                                        # Ini Total Section (sudah termasuk PPN/diskon)
                                        result['sections'][section_letter]['total_row'] = row
                                        result['sections'][section_letter]['total_value'] = total_value
                                    else:
                                        # Ini Jumlah/Subtotal (belum PPN)
                                        result['sections'][section_letter]['subtotal_row'] = row
                                        result['sections'][section_letter]['subtotal_value'] = total_value
                                
                                current_section = section_letter
                                
                                if result['subtotal_value'] is None:
                                    result['subtotal_row'] = row
                                    result['subtotal_value'] = total_value
                            else:
                                # Generic total tanpa label section
                                if current_section:
                                    # Check Total vs Jumlah
                                    cell_upper = cell_str.upper()
                                    if 'TOTAL' in cell_upper and 'JUMLAH' not in cell_upper:
                                        result['sections'][current_section]['total_row'] = row
                                        result['sections'][current_section]['total_value'] = total_value
                                    else:
                                        result['sections'][current_section]['subtotal_row'] = row
                                        result['sections'][current_section]['subtotal_value'] = total_value
                                else:
                                    # Belum ada section, buat default
                                    result['sections']['A'] = {
                                        'items': [],
                                        'subtotal_row': row,
                                        'subtotal_value': total_value,
                                        'ppn_row': None,
                                        'ppn_value': None,
                                        'discount_row': None,
                                        'discount_value': None,
                                        'total_row': None,
                                        'total_value': None
                                    }
                                    current_section = 'A'
                                    section_order.append('A')
                                
                                if result['subtotal_value'] is None:
                                    result['subtotal_row'] = row
                                    result['subtotal_value'] = total_value
                            
                            is_summary_row = True
                            break
            
            # Jika bukan baris summary, baca sebagai item
            if not is_summary_row:
                item = self.read_item(row, columns)
                
                unit_price = item.get('unit_price')
                total = item.get('total')
                
                if isinstance(unit_price, str) and ('Rp' in unit_price or 'Rp.' in unit_price):
                    continue
                if isinstance(total, str) and ('Rp' in total or 'Rp.' in total):
                    continue
                
                has_valid_data = False
                row_dump = None
                
                # STEP 1: always compute total as qty * unit_price as source of truth
                qty = safe_float(item.get('qty'))
                up = safe_float(item.get('unit_price'))
                excel_total = safe_float(item.get('total'))
                
                calc_total = None
                if qty is not None and up is not None:
                    calc_total = qty * up
                
                # STEP 2: decide item total = calc_total if available, else excel_total
                if calc_total is not None:
                    item['total'] = calc_total
                    # Keep raw excel total for error detection if different
                    if excel_total is not None and abs(excel_total - calc_total) > 0.01:
                        item['excel_total_raw'] = excel_total
                        item['calc_mismatch'] = True
                    has_valid_data = True
                elif excel_total is not None:
                    item['total'] = excel_total
                    has_valid_data = True
                elif qty is not None:
                    # Qty exists but no price — keep item so debug is visible
                    has_valid_data = True
                else:
                    # No valid data — log for debug why row skipped
                    row_dump = [str(self.ws.cell(row=row, column=c).value)[:12] if self.ws.cell(row=row, column=c).value else "" for c in range(1, 8)]
                    skipped_rows.append({'row': row, 'dump': row_dump})
                
                if has_valid_data:
                    result['items'].append(item)
                    
                    # Selalu tambahkan ke section aktif atau pending
                    if current_section and current_section in result['sections']:
                        result['sections'][current_section]['items'].append(item)
                    else:
                        pending_items.append(item)
        
        # Handle pending items - PENTING: pastikan semua item masuk ke section
        if pending_items:
            if not result['sections']:
                # Belum ada section, buat section A
                result['sections']['A'] = {
                    'items': pending_items,
                    'subtotal_row': None,
                    'subtotal_value': None,
                    'ppn_row': None,
                    'ppn_value': None,
                    'discount_row': None,
                    'discount_value': None,
                    'total_row': None,
                    'total_value': None
                }
            else:
                # Sudah ada section, masukkan ke section pertama
                first_section = min(result['sections'].keys())
                result['sections'][first_section]['items'].extend(pending_items)
        
        # TOTAL (A+B) di atas PPN vs GRAND TOTAL setelah PPN — jangan promosikan TOTAL (A+B) yang sebelum PPN jadi grand
        # Hanya grand yang setelah PPN yang grand. TOTAL (A+B) sebelum PPN adalah jumlah_global_excel/total_kategori_value
        pass

        # Hitung subtotal per section jika belum ada — TETAP tampil Jumlah A/B sebelum PPN
        # Case Sparepart/Instalasi: A/B hanya pemisah kategori, bukan section dijumlah terpisah.
        # TOTAL tanpa huruf (Row 28) = Jumlah Global sebelum PPN -> pisahkan dari Grand agar tidak dobel.
        is_labeled_section = any(v.get('subtotal_value') is not None and not v.get('subtotal_is_calculated', True)
                                 or v.get('total_value') is not None for v in result['sections'].values())
        # Jika tidak ada Total/Jumlah per-section dari Excel (hanya A/B kategori), TOTAL global adalah Jumlah Global, bukan Grand
        # Tandai agar checker tidak cek section sebagai error dan kategori tetap tampil Jumlah A/B
        for section_letter, section_data in result['sections'].items():
            if section_data['subtotal_value'] is None:
                calc_subtotal = 0
                for i in section_data['items']:
                    val = safe_float(i.get('total'))
                    if val is not None:
                        calc_subtotal += val
                section_data['subtotal_value'] = calc_subtotal
                section_data['subtotal_is_calculated'] = True
            else:
                section_data['subtotal_is_calculated'] = False
            # Kategori kategori: jika belum ada Jumlah per-section dari Excel, flag kategori agar tidak dianggap section terpisah untuk total
            section_data['is_category'] = not is_labeled_section
        
        # Jumlah Global: jika ada Jumlah Global eksplisit, jangan isi subtotal_value global dengan Total A
        # (sudah dipisah via jumlah_global_excel). Untuk tanpa-Jumlah kasus, fallback SUM item hanya jika tanpa PPN / single.
        if result['subtotal_value'] is None and len(result['sections']) <= 1:
            result['subtotal_value'] = 0
            for i in result['items']:
                val = safe_float(i.get('total'))
                if val is not None:
                    result['subtotal_value'] += val
        
        # === KLASIFIKASI PPN setelah semua subtotal/total terbaca ===
        ppn_mode = overrides.get('ppn_mode', 'auto')
        if ppn_candidates:
            if ppn_mode == 'combined':
                # User memaksa: satu PPN gabungan untuk semua section
                sum_sub_all = sum(safe_float(v.get('subtotal_value')) or 0 for v in result['sections'].values())
                best = max(ppn_candidates, key=lambda c: c['row'])  # ambil yang paling bawah (biasanya gabungan)
                result['ppn_value'] = best['value']
                result['ppn_row'] = best['row']
                result['ppn_is_combined'] = True
            elif ppn_mode == 'per_section':
                # User memaksa: semua generik adalah per-section
                section_order_sorted = sorted(result['sections'].keys())
                for cand in ppn_candidates:
                    sec_letter = self._detect_section_letter(cand['label'])
                    if sec_letter and sec_letter != 'GRAND' and sec_letter in result['sections']:
                        result['sections'][sec_letter]['ppn_value'] = cand['value']
                        result['sections'][sec_letter]['ppn_row'] = cand['row']
                    else:
                        for sl in section_order_sorted:
                            if result['sections'][sl]['ppn_value'] is None:
                                result['sections'][sl]['ppn_value'] = cand['value']
                                result['sections'][sl]['ppn_row'] = cand['row']
                                break
                # also set global as sum for display
                s = sum(safe_float(v.get('ppn_value')) or 0 for v in result['sections'].values())
                if s > 0:
                    result['ppn_value'] = s
                    result['ppn_is_combined'] = False
            elif ppn_mode == 'single':
                # Hanya section tertentu (paling bawah / yang punya PPN saja)
                best = max(ppn_candidates, key=lambda c: c['row'])
                pref = best['current_section']
                target = pref if pref and pref in result['sections'] else sorted(result['sections'].keys())[-1]
                result['sections'][target]['ppn_value'] = best['value']
                result['sections'][target]['ppn_row'] = best['row']
                result['ppn_value'] = best['value']
                result['ppn_row'] = best['row']
                result['ppn_is_combined'] = False
            elif ppn_mode == 'none':
                pass  # jangan isi PPN sama sekali
            else:  # auto — PPN urutan tetap: A, B, lalu TOTAL/GRAND. Jika 1 PPN gabungan setelah TOTAL, jangan masuk section
                # Langkah 1: explicit label PPN ... A/B -> langsung
                remaining = []
                for cand in ppn_candidates:
                    sec_letter = self._detect_section_letter(cand['label'])
                    if sec_letter and sec_letter != 'GRAND' and sec_letter in result['sections']:
                        result['sections'][sec_letter]['ppn_value'] = cand['value']
                        result['sections'][sec_letter]['ppn_row'] = cand['row']
                    else:
                        remaining.append(cand)
                if remaining:
                    # Langkah 2: deteksi PPN GABUNGAN dulu (setelah Jumlah A+B) — JANGAN masuk ke A
                    # PPN gabungan = 11% * (Jumlah A + Jumlah B) = 11% * jumlah_global_excel atau sum_sub_all
                    jumlah_global_for_ppn = safe_float(result.get('jumlah_global_excel'))
                    if jumlah_global_for_ppn is None:
                        jumlah_global_for_ppn = sum(safe_float(v.get('subtotal_value')) or 0 for v in result['sections'].values())
                    combined_candidate = None
                    # PPN 1 BAGIAN (hanya B punya PPN) tidak boleh dianggap gabungan meski nilainya dekat TOTAL×11% secara kebetulan
                    # Gabungan = 11% * Jumlah Global (A+B), sedangkan PPN B = 11% * Jumlah B — keduanya berbeda jika ada A
                    # Jadi gabungan hanya jika PPN tidak dekat dengan salah satu Jumlah section
                    if len(remaining) >= 1 and jumlah_global_for_ppn and jumlah_global_for_ppn > 0:
                        expected_combined = jumlah_global_for_ppn * 0.11
                        for cand in remaining:
                            cand_val = safe_float(cand.get('value'))
                            if cand_val is None:
                                continue
                            # Cek bukan PPN per-section-nya B (11% * Jumlah B)
                            is_ppn_per_section = False
                            for sl, sd in result['sections'].items():
                                sub = safe_float(sd.get('subtotal_value'))
                                if sub and abs(cand_val - sub * 0.11) <= max(2, sub * 0.11 * 0.008):
                                    is_ppn_per_section = True
                                    break
                            if is_ppn_per_section:
                                continue
                            if abs(cand_val - expected_combined) <= max(2, expected_combined * 0.008):
                                combined_candidate = cand
                                break
                    if combined_candidate is not None:
                        result['ppn_value'] = combined_candidate['value']
                        result['ppn_row'] = combined_candidate['row']
                        result['ppn_is_combined'] = True
                        remaining = [c for c in remaining if c is not combined_candidate]
                    # Langkah 3: sisa baru distribusikan per-section (urut A->B)
                    section_order_sorted = sorted(result['sections'].keys())
                    for cand in remaining:
                        placed = False
                        pref = cand['current_section']
                        if pref and pref in result['sections'] and result['sections'][pref]['ppn_value'] is None:
                            # Hanya isi section jika belum ada tanda PPN gabungan global
                            if not result.get('ppn_is_combined'):
                                result['sections'][pref]['ppn_value'] = cand['value']
                                result['sections'][pref]['ppn_row'] = cand['row']
                                placed = True
                            else:
                                # Sudah ada gabungan, lebih aman simpan sebagai global juga
                                if result.get('ppn_value') is None:
                                    result['ppn_value'] = cand['value']
                                    result['ppn_row'] = cand['row']
                                placed = True
                        # Fallback global jika tidak ada yang match PPN per-section
                        for sl in section_order_sorted:
                            if result['sections'][sl]['ppn_value'] is None:
                                sub_b = safe_float(result['sections'][sl].get('subtotal_value'))
                                cand_ppn = safe_float(cand.get('value'))
                                if sub_b and cand_ppn and abs(cand_ppn - sub_b * 0.11) > max(2, sub_b * 0.11 * 0.015):
                                    continue
                                result['sections'][sl]['ppn_value'] = cand['value']
                                result['sections'][sl]['ppn_row'] = cand['row']
                                placed = True
                                break
                        if not placed and result.get('ppn_value') is None:
                            result['ppn_value'] = cand['value']
                            result['ppn_row'] = cand['row']
                            result['ppn_is_combined'] = True
                    if len(result['sections']) == 1 and result.get('ppn_value') is None:
                        pass

        # Expose jumlah section dinamis (A/B/C atau lebih) untuk laporan
        result['detected_sections'] = sorted(result['sections'].keys())

        # CASE 1: RAB TANPA PPN — jika tidak ada PPN di mana pun, GRAND TOTAL = TOTAL = SUM item
        if result.get('ppn_value') is None and all(safe_float(v.get('ppn_value')) is None for v in result['sections'].values()):
            result['is_without_ppn'] = True
            # Grand = TOTAL gabungan jika ada, else Jumlah Global / subtotal / SUM item langsung
            if result['grand_total_value'] is None:
                cand = result.get('jumlah_global_excel') or result.get('subtotal_value')
                if cand is not None:
                    result['grand_total_value'] = safe_float(cand)
                    result['grand_total_row'] = result.get('jumlah_global_row') or result.get('subtotal_row')
                elif len(result['sections']) > 1:
                    s = sum(safe_float(v.get('subtotal_value')) or 0 for v in result['sections'].values())
                    if not s:
                        s = sum(safe_float(it.get('total')) or 0 for it in result['items'])
                    if s:
                        result['grand_total_value'] = s
                        result['grand_total_row'] = result.get('subtotal_row') or result.get('jumlah_global_row')
                elif result['sections']:
                    only = list(result['sections'].values())[0]
                    if safe_float(only.get('subtotal_value')) is not None:
                        result['grand_total_value'] = safe_float(only.get('subtotal_value'))
                        result['grand_total_row'] = only.get('subtotal_row')
                    else:
                        # Fallback paling simpel: SUM semua item (CASE 1 tanpa PPN paling dasar)
                        s = sum(safe_float(it.get('total')) or 0 for it in result['items'])
                        if s:
                            result['grand_total_value'] = s
                            result['subtotal_value'] = s
                            result['jumlah_global_excel'] = s
                            only['subtotal_value'] = s
                else:
                    # Tanpa section sama sekali (file simpel Item->TOTAL)
                    s = sum(safe_float(it.get('total')) or 0 for it in result['items'])
                    if s:
                        result['grand_total_value'] = s
                        result['subtotal_value'] = s
                        result['jumlah_global_excel'] = s
            # Pastikan subtotal/grand tidak tetap None untuk case tanpa ppn paling mudah
            if result.get('subtotal_value') is None:
                s = sum(safe_float(it.get('total')) or 0 for it in result['items'])
                if s:
                    result['subtotal_value'] = s
                    result['jumlah_global_excel'] = s
            if not result['sections'] and result['grand_total_value'] is None and result.get('subtotal_value') is not None:
                result['grand_total_value'] = safe_float(result.get('subtotal_value'))
        else:
            result['is_without_ppn'] = False

        # === VALUE INTELLIGENCE: reconciliator Jumlah Global vs Grand Total via angka ===
        # Jika GRAND TOTAL masih kosong tapi Jumlah Global ada + PPN gabungan, derived Grand = Jumlah Global + PPN
        # Jika Jumlah Global kosong tapi Grand ada, jangan tebak — hanya jika TOTAL tunggal case.
        # Ini tetap alur RAB (angka sebagai kebenaran), tulisan hanya petunjuk.
        if result['grand_total_value'] is None and result.get('jumlah_global_excel') is not None:
            jg = safe_float(result.get('jumlah_global_excel'))
            ppn_g = safe_float(result.get('ppn_value')) if result.get('ppn_is_combined') else None
            if jg is not None and ppn_g is not None:
                # Jangan override jika sudah ada grand via sections; ini hanya jika sections belum punya grand
                if result['grand_total_value'] is None:
                    pass  # Grand tetap dari sections jika ada; Quantity tetap Jumlah Global untuk check_global_subtotal
        # Jika PPN gabungan 10.184.363 tapi GRAND belum match (mis. Excel GRAND sudah termasuk PPN), papan di app akan reconcile via sum
            
        # Fallback global: jika masih tanpa nilai setelah CASE 1, aseg SUM item (jumlah sebelum PPN)
        # Kasus tanpa section & tanpa TOTAL label: pastikan tidak tetap Rp 0
        if result.get('subtotal_value') is None or result.get('subtotal_value') == 0:
            s_items = sum(safe_float(it.get('total')) or 0 for it in result['items'])
            if s_items:
                result['subtotal_value'] = s_items
                result['jumlah_global_excel'] = s_items
                result['jumlah_global_row'] = result.get('jumlah_global_row') or result.get('subtotal_row') or result.get('header_row')
                # Isi ke classifications debug agar tampil value
                if not any(c.get('type') == 'jumlah_global' for c in classifications):
                    classifications.append({'row': result.get('header_row') or 0, 'raw': 'TOTAL (auto SUM)', 'normalized': 'TOTAL', 'type': 'jumlah_global', 'fuzzy': False})
        if result.get('grand_total_value') is None and result.get('is_without_ppn') and result.get('subtotal_value'):
            result['grand_total_value'] = safe_float(result.get('subtotal_value'))
            result['grand_total_row'] = result.get('subtotal_row') or result.get('jumlah_global_row')

        # Hitung grand total dari sections jika belum ada (fleksibel)
        if result['grand_total_value'] is None and len(result['sections']) > 1:
            total_all = 0
            for sl, sd in result['sections'].items():
                sec_total = safe_float(sd.get('total_value'))
                if sec_total is None:
                    sub = safe_float(sd.get('subtotal_value')) or 0
                    ppn = safe_float(sd.get('ppn_value')) or 0
                    disc = safe_float(sd.get('discount_value')) or 0
                    sec_total = sub + ppn - disc if (sub or ppn or disc) else None
                if sec_total is not None:
                    total_all += sec_total
            if total_all > 0:
                result['grand_total_value'] = total_all
        # Grand total sudah ada tapi PPN gabungan belum terdeteksi: jangan timpa
        
            # PPN 1 BAGIAN: jangan sum ke global (itu gabungan palsu)
            # Hanya sum jika PPN per-section murni (lebih dari 1 section punya PPN)
            ppn_sections_cnt = sum(1 for v in result['sections'].values() if safe_float(v.get('ppn_value')) is not None)
            if result.get('ppn_value') is None and ppn_sections_cnt > 1:
                ppn_sum = sum(safe_float(v.get('ppn_value')) or 0 for v in result['sections'].values())
                if ppn_sum > 0:
                    result['ppn_value'] = ppn_sum
                    result['ppn_is_combined'] = False
        result.setdefault('ppn_is_combined', False)
        result.setdefault('is_without_ppn', False)

        result['skipped_rows'] = skipped_rows
        result['classifications'] = classifications
        result['columns'] = columns
        result['summary_rows_debug'] = []
        for k in classifications:
            try:
                v_dbg = self._get_total_value(k['row'], total_col)
                if v_dbg is None:
                    for c2 in range(self.ws.max_column, 0, -1):
                        v2 = self.ws_data.cell(row=k['row'], column=c2).value
                        if v2 is None:
                            v2 = self.ws.cell(row=k['row'], column=c2).value
                        sf = safe_float(v2)
                        if sf is not None and sf > 0:
                            v_dbg = sf
                            break
                k2 = dict(k)
                k2['value'] = v_dbg
                result['summary_rows_debug'].append(k2)
            except Exception:
                pass
        # AI opsional — tanpa AI fallback lokal gratis (Value Intelligence)
        if ai_overrides and isinstance(ai_overrides, dict) and ai_overrides.get('provider') != 'none':
            try:
                should_ai = isinstance(ai_overrides, dict) and ai_overrides.get('provider') != 'none'
                if should_ai and not (ai_overrides.get('gemini_key') or ai_overrides.get('groq_key')):
                    import os as _os
                    if _os.environ.get("GOOGLE_API_KEY") or _os.environ.get("GEMINI_API_KEY"):
                        ai_overrides = {'provider': 'gemini', 'gemini_key': _os.environ.get("GOOGLE_API_KEY") or _os.environ.get("GEMINI_API_KEY")}
                    else:
                        should_ai = False
                if should_ai:
                    ai_map = None
                    tmp_summary = []
                    for k in classifications:
                        try:
                            v_dbg2 = None
                            for c2 in range(self.ws.max_column, 0, -1):
                                v2 = self.ws_data.cell(row=k['row'], column=c2).value
                                if v2 is None:
                                    v2 = self.ws.cell(row=k['row'], column=c2).value
                                sf = safe_float(v2)
                                if sf is not None and sf > 0:
                                    v_dbg2 = sf
                                    break
                        except: v_dbg2 = None
                        tmp_summary.append({'row': k['row'], 'value': v_dbg2})
                    klass_map_tmp = {x['row']: x for x in tmp_summary}
                    g_key2 = (ai_overrides or {}).get('gemini_key') or (ai_overrides or {}).get('groq_key')
                    if (ai_overrides or {}).get('provider') == 'gemini' or (g_key2 and 'gemini' in str((ai_overrides or {}).get('provider','')).lower()):
                        from ai_helper import classify_with_gemini_free
                        rows_for_ai = [{'row': k['row'], 'raw': k['raw'], 'normalized': k.get('normalized',''), 'value': klass_map_tmp.get(k['row'],{}).get('value')} for k in classifications]
                        ai_map = classify_with_gemini_free(rows_for_ai, api_key=(ai_overrides or {}).get('gemini_key'))
                    elif (ai_overrides or {}).get('provider') == 'groq':
                        from ai_helper import classify_with_groq_free
                        rows_for_ai = [{'row': k['row'], 'raw': k['raw']} for k in classifications]
                        ai_map = classify_with_groq_free(rows_for_ai, api_key=(ai_overrides or {}).get('groq_key'))
                    if ai_map:
                        for k in result['classifications']:
                            if k['row'] in ai_map and ai_map[k['row']] in ('jumlah_global','ppn','grand_total','subtotal','discount','unknown','section_header'):
                                k['type'] = ai_map[k['row']]
                                k['fuzzy'] = True
                                k['ai_patched'] = True
            except Exception:
                pass
        return result
    
    def read_item(self, row: int, columns: Dict[str, int]) -> Dict[str, Any]:
        """Membaca satu baris item dengan formula dan nilai"""
        item = {
            'row': row,
            'item_name': None,
            'description': None,
            'qty': None,
            'harga_awal': None,
            'mark_up': None,
            'unit_price': None,
            'total': None,
            'total_formula': None,
            'has_formula': False
        }
        
        if columns.get('item_name'):
            cell = self.ws.cell(row=row, column=columns['item_name'])
            item['item_name'] = cell.value
            item['description'] = cell.value
        
        if not item['item_name'] or (isinstance(item['item_name'], str) and len(item['item_name']) > 30):
            for col in range(1, columns.get('qty', 6)):
                cell = self.ws.cell(row=row, column=col)
                cell_value = cell.value
                if cell_value and isinstance(cell_value, str):
                    if cell_value.isdigit() or cell_value.startswith('=') or len(cell_value) < 2:
                        continue
                    if cell_value.upper() in ['RP.', 'RP', 'NO.', 'NO', 'UNIT', 'QTY', 'DESCRIPTION', 'ITEM', 'NAME']:
                        continue
                    item['item_name'] = cell_value
                    item['description'] = cell_value
                    break
        
        if columns.get('qty'):
            cell_data = self.ws_data.cell(row=row, column=columns['qty'])
            cell_formula = self.ws.cell(row=row, column=columns['qty'])
            raw_val = cell_data.value if cell_data.value is not None else cell_formula.value
            item['qty_raw'] = raw_val
            item['qty'] = safe_float(raw_val)
        
        if columns.get('harga_awal'):
            cell_data = self.ws_data.cell(row=row, column=columns['harga_awal'])
            cell_formula = self.ws.cell(row=row, column=columns['harga_awal'])
            item['harga_awal'] = cell_data.value if cell_data.value is not None else cell_formula.value
        
        if columns.get('mark_up'):
            cell_data = self.ws_data.cell(row=row, column=columns['mark_up'])
            cell_formula = self.ws.cell(row=row, column=columns['mark_up'])
            item['mark_up'] = cell_data.value if cell_data.value is not None else cell_formula.value
        
        if columns.get('unit_price'):
            cell_data = self.ws_data.cell(row=row, column=columns['unit_price'])
            cell_formula = self.ws.cell(row=row, column=columns['unit_price'])
            raw_val = cell_data.value if cell_data.value is not None else cell_formula.value
            item['unit_price_raw'] = raw_val
            if raw_val is not None:
                converted = safe_float(raw_val)
                if converted is not None:
                    item['unit_price'] = converted
                elif cell_formula.data_type == 'f':
                    item['has_formula'] = True
                    item['unit_price'] = None
                else:
                    item['unit_price'] = raw_val
        
        if columns.get('total'):
            cell_data = self.ws_data.cell(row=row, column=columns['total'])
            cell_formula = self.ws.cell(row=row, column=columns['total'])
            
            raw_val = cell_data.value if cell_data.value is not None else cell_formula.value
            item['total_raw'] = raw_val
            item['total_formula_exposed'] = cell_formula.value if cell_formula.data_type == 'f' else None
            if raw_val is not None:
                converted = safe_float(raw_val)
                if converted is not None:
                    item['total'] = converted
                elif cell_formula.data_type == 'f':
                    item['has_formula'] = True
                    item['total_formula'] = cell_formula.value
                    if item.get('qty') is not None and item.get('unit_price') is not None:
                        item['total'] = item['qty'] * item['unit_price']
                else:
                    item['total'] = raw_val
        
        return item
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """Mendapatkan nilai cell"""
        if self.ws is None:
            return None
        return self.ws.cell(row=row, column=col).value
    
    def get_cell_formula(self, row: int, col: int) -> Optional[str]:
        """Mendapatkan formula cell"""
        if self.ws is None:
            return None
        cell = self.ws.cell(row=row, column=col)
        if cell.data_type == 'f':
            return cell.value
        return None
